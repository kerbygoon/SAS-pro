import os, base64
import streamlit as st
from streamlit_autorefresh import st_autorefresh
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd

# ===== Paths =====
base_folder = r"murid"
excel_file = r"test.xlsx"

# ===== Function to set background =====
def add_bg_from_local(image_file):
    with open(image_file, "rb") as f:
        encoded = base64.b64encode(f.read()).decode()
    st.markdown(
        f"""
        <style>
        [data-testid="stAppViewContainer"] {{
            background-image: url("data:image/png;base64,{encoded}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
        }}
        [data-testid="stHeader"] {{
            background: rgba(0,0,0,0);
        }}
        html, body, [class*="css"] {{
            color: white;
        }}
        h1, h2, h3, h4, h5, h6, p, label, div {{
            color: white !important;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

add_bg_from_local(r"bgggggggg.png")  # change if needed

# ===== Auto-generate student list =====
paths = {
    os.path.splitext(filename)[0]: os.path.join(base_folder, filename)
    for filename in os.listdir(base_folder)
    if filename.endswith(".text")
}
students = sorted(paths.keys())

# ===== Initialize session state =====
for name in students:
    if name not in st.session_state:
        st.session_state[name] = False

st.subheader("📚 Class: 2 Alpha - Attendance System")

# ===== Helper Functions =====
def check_att(path, name):
    if os.path.exists(path):
        with open(path, "r") as f:
            first_val = f.read().strip().split(",")[0].strip()
            st.session_state[name] = (first_val == "True")

def save_to_excel(present, absent, reasons, date_str):
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Student", "Date", "Status", "Cause"])
    else:
        wb = load_workbook(excel_file)
        ws = wb.active

    for student in present:
        ws.append([student, date_str, "Present", ""])

    for student in absent:
        ws.append([student, date_str, "Absent", reasons.get(student, "")])

    wb.save(excel_file)

def analyze_attendance():
    if not os.path.exists(excel_file):
        st.warning("⚠️ No attendance records found yet!")
        return

    df = pd.read_excel(excel_file)
    df.columns = ["Student", "Date", "Status", "Cause"]
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.sort_values(by=["Student", "Date"])

    def longest_absent_streak(student_df):
        student_df = student_df.sort_values("Date")
        student_df["AbsentFlag"] = student_df["Status"] == "Absent"
        student_df["DayGap"] = student_df["Date"].diff().dt.days.fillna(0)

        max_streak = 0
        current_streak = 0
        last_date = None

        for _, row in student_df.iterrows():
            if row["AbsentFlag"]:
                if last_date is None or row["DayGap"] == 1:
                    current_streak += 1
                else:
                    current_streak = 1
                max_streak = max(max_streak, current_streak)
            else:
                current_streak = 0
            last_date = row["Date"]

        return max_streak

    results = []

    for student in df["Student"].unique():
        student_df = df[df["Student"] == student]

        total_absences = (student_df["Status"] == "Absent").sum()
        total_presents = (student_df["Status"] == "Present").sum()
        max_streak = longest_absent_streak(student_df)

        if max_streak >= 31 or total_absences >= 60:
            level = "🚨 Expulsion"
        elif max_streak >= 17 or total_absences >= 40:
            level = "⚠️ Last Note"
        elif max_streak >= 10 or total_absences >= 20:
            level = "⚠️ Second Note"
        elif max_streak >= 3 or total_absences >= 10:
            level = "⚠️ First Note"
        else:
            level = "✅ Good Standing"

        absent_info = student_df[student_df["Status"] == "Absent"][["Date", "Cause"]]
        absent_details = [
            f"{row.Date.strftime('%Y-%m-%d')} ({row.Cause})"
            for row in absent_info.itertuples()
        ]

        results.append([
            student,
            total_presents,
            total_absences,
            max_streak,
            level,
            ", ".join(absent_details)
        ])

    results_df = pd.DataFrame(
        results,
        columns=[
            "Student",
            "Total Present Days",
            "Total Absences",
            "Max Consecutive Absences",
            "Status",
            "Absent Dates & Causes"
        ]
    )

    st.dataframe(results_df, use_container_width=True)
# ===== Reset Attendance =====
if st.button("🔄 Reset All Attendance"):
    for name, path in paths.items():
        with open(path, "w") as f:
            f.write("False,")
        st.session_state[name] = False
    st.success("✅ All attendance reset.")

# ===== Auto-refresh =====
st_autorefresh(interval=30000, limit=None, key="attendance_refresh")

# ===== Read attendance from files =====
for name, path in paths.items():
    check_att(path, name)

# ===== Date Selection =====
st.write("### 📅 Attendance Date")
selected_date = st.date_input(
    "Select attendance date",
    value=datetime.now().date()
)
selected_date_str = selected_date.strftime("%Y-%m-%d")

# ===== Attendance Checkboxes =====
st.write("### ✅ Mark Attendance")
for name in students:
    st.session_state[name] = st.checkbox(
        name.capitalize(),
        value=st.session_state[name]
    )

# ===== Absence Reasons =====
absence_reasons = {}
st.write("### 📝 Cause of Absence (Absent Students)")
for name in students:
    if not st.session_state[name]:
        absence_reasons[name] = st.text_input(
            f"Reason for {name.capitalize()}",
            placeholder="Sick / Family / No reason"
        )

# ===== Submit Attendance ONLY =====
if st.button("📥 Submit Attendance"):
    present = [name for name in students if st.session_state[name]]
    absent = [name for name in students if not st.session_state[name]]

    save_to_excel(present, absent, absence_reasons, selected_date_str)

    st.success("✅ Attendance recorded!")
    st.write("**Present:**", ", ".join(present) if present else "None")
    st.write("**Absent:**", ", ".join(absent) if absent else "None")

# ===== Show Analysis ONLY =====
if st.button("📊 Show Attendance Analysis"):
    st.subheader("📊 Attendance Analysis")
    analyze_attendance()

